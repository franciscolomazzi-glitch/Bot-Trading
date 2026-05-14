import ccxt
import config
import time
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os

# Conexión a Binance (solo lectura de precios)
exchange = ccxt.binance()

# Archivo Excel
ARCHIVO_EXCEL = "operaciones.xlsx"

def inicializar_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Operaciones"
        ws.append(["Fecha", "Hora", "Tipo", "Precio", "Capital", "Ganancia bruta", "Comisión (0.09%)", "Ganancia neta", "Acumulado"])
        wb.save(ARCHIVO_EXCEL)

def registrar_operacion(tipo, precio, ganancia_bruta, acumulado):
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    ahora = datetime.now()
    comision = precio * 0.0009
    ganancia_neta = ganancia_bruta - comision
    ws.append([
        ahora.strftime("%d/%m/%Y"),
        ahora.strftime("%H:%M:%S"),
        tipo,
        precio,
        config.TOTAL_CAPITAL / config.GRID_LEVELS,
        round(ganancia_bruta, 4),
        round(comision, 4),
        round(ganancia_neta, 4),
        round(acumulado, 4)
    ])
    wb.save(ARCHIVO_EXCEL)

def obtener_precio_actual():
    ticker = exchange.fetch_ticker(config.SYMBOL)
    return ticker['last']

def calcular_grilla():
    niveles = []
    rango = config.UPPER_PRICE - config.LOWER_PRICE
    paso = rango / config.GRID_LEVELS
    for i in range(config.GRID_LEVELS + 1):
        precio = config.LOWER_PRICE + (paso * i)
        niveles.append(round(precio, 2))
    return niveles

def ejecutar_bot():
    inicializar_excel()
    print("🤖 Bot iniciado en modo PAPER TRADING (sin dinero real)")
    print(f"Par: {config.SYMBOL}")
    grilla = calcular_grilla()
    print(f"Niveles de grilla: {grilla}")
    capital_por_nivel = config.TOTAL_CAPITAL / config.GRID_LEVELS
    print(f"Capital por nivel: {capital_por_nivel} USDT")
    print("─" * 50)

    ultimo_precio = None
    ganancias = 0

    while True:
        try:
            precio_actual = obtener_precio_actual()
            print(f"\n💲 Precio actual BTC: ${precio_actual:,.2f}")

            for i in range(len(grilla) - 1):
                nivel_compra = grilla[i]
                nivel_venta = grilla[i + 1]
                ganancia_celda = (nivel_venta - nivel_compra) / nivel_compra * capital_por_nivel

                if ultimo_precio and ultimo_precio > nivel_compra and precio_actual <= nivel_compra:
                    print(f"🟢 COMPRA simulada en ${nivel_compra:,.2f}")
                    registrar_operacion("COMPRA", nivel_compra, 0, ganancias)

                if ultimo_precio and ultimo_precio < nivel_venta and precio_actual >= nivel_venta:
                    ganancias += ganancia_celda
                    print(f"🔴 VENTA simulada en ${nivel_venta:,.2f} | Ganancia celda: ${ganancia_celda:.2f} | Total acumulado: ${ganancias:.2f}")
                    registrar_operacion("VENTA", nivel_venta, ganancia_celda, ganancias)

            ultimo_precio = precio_actual
            time.sleep(10)

        except Exception as e:
            print(f"Error: {e}")
            time.sleep(10)

ejecutar_bot()